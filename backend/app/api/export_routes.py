"""Analytics and export. Plan sections 8 and 10; Phase 5.

Exports are posting-level, not transaction-level. A transaction has no single
amount -- that is the whole point of the double-entry model -- so a row-per-
transaction CSV would have to invent one, and any invented figure is the one
that stops reconciling.
"""

from __future__ import annotations

import csv
import io
import json
import uuid
from datetime import date
from decimal import Decimal

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.api.schemas import to_minor
from app.db import get_session
from app.domain import analytics
from app.domain import backup as backup_module
from app.domain import restore as restore_module
from app.domain.clock import today as clock_today
from app.domain.ledger_scope import posted_transaction_ids
from app.models import Account, Category, Posting, Transaction
from app.models.enums import LIQUID_KINDS, AccountKind

router = APIRouter()


class CategoryTotalOut(BaseModel):
    #: None for the "Uncategorised" row -- there is no category to link to.
    category_id: uuid.UUID | None
    name: str
    amount_minor: int


class PeriodSummaryOut(BaseModel):
    start: date
    end: date
    income_minor: int
    expense_minor: int
    saved_minor: int
    net_minor: int
    #: (income - spending) / income. None when there was no income.
    savings_rate: float | None
    #: Deliberately moved to savings or investments, as a share of income.
    set_aside_rate: float | None
    by_category: list[CategoryTotalOut]
    by_merchant: list[tuple[str, int]]


def _summary_out(s: analytics.PeriodSummary) -> PeriodSummaryOut:
    return PeriodSummaryOut(
        start=s.start,
        end=s.end,
        income_minor=to_minor(s.income),
        expense_minor=to_minor(s.expense),
        saved_minor=to_minor(s.saved),
        net_minor=to_minor(s.net),
        savings_rate=float(s.savings_rate) if s.savings_rate is not None else None,
        set_aside_rate=float(s.set_aside_rate) if s.set_aside_rate is not None else None,
        by_category=[
            CategoryTotalOut(
                category_id=c.category_id, name=c.name, amount_minor=to_minor(c.amount)
            )
            for c in s.by_category
        ],
        by_merchant=[(name, to_minor(total)) for name, total in s.by_merchant],
    )


@router.get("/analytics/period", response_model=PeriodSummaryOut)
def period_summary(
    start: date | None = None,
    end: date | None = None,
    session: Session = Depends(get_session),
) -> PeriodSummaryOut:
    today = clock_today(session)
    month_start, month_end = analytics.month_bounds(today.year, today.month)
    start = start or month_start
    end = end or month_end
    return _summary_out(analytics.summarise(session, start, end))


@router.get("/analytics/monthly", response_model=list[PeriodSummaryOut])
def monthly(
    first: date | None = None,
    last: date | None = None,
    session: Session = Depends(get_session),
) -> list[PeriodSummaryOut]:
    today = clock_today(session)
    last = last or today
    first = first or date(last.year, 1, 1)
    return [_summary_out(s) for s in analytics.monthly_series(session, first, last)]


# --------------------------------------------------------------------------
# Export
# --------------------------------------------------------------------------


def _rows(session: Session, start: date, end: date):
    """One row per posting, joined to its transaction and account."""
    return session.execute(
        select(Transaction, Posting, Account, Category)
        .join(Posting, Posting.transaction_id == Transaction.id)
        .join(Account, Posting.account_id == Account.id)
        .outerjoin(Category, Posting.category_id == Category.id)
        .where(Posting.transaction_id.in_(posted_transaction_ids(start=start, end=end)))
        .order_by(Transaction.booking_date, Transaction.id)
    ).all()


@router.get("/export/transactions.csv")
def export_csv(
    start: date | None = None,
    end: date | None = None,
    session: Session = Depends(get_session),
) -> StreamingResponse:
    """Posting-level CSV. Amounts are decimal strings, never floats."""
    today = clock_today(session)
    start = start or date(today.year, 1, 1)
    end = end or today

    buffer = io.StringIO()
    writer = csv.writer(buffer)
    writer.writerow(
        [
            "booking_date",
            "transaction_id",
            "description",
            "merchant",
            "account",
            "account_kind",
            "category",
            "amount",
            "currency",
        ]
    )
    for txn, posting, account, category in _rows(session, start, end):
        writer.writerow(
            [
                txn.booking_date.isoformat(),
                str(txn.id),
                txn.description,
                txn.merchant or "",
                account.name,
                account.kind.value,
                category.name if category else "",
                # str(Decimal) keeps the exact scale; float() would not.
                str(posting.amount),
                posting.currency,
            ]
        )

    buffer.seek(0)
    return StreamingResponse(
        iter([buffer.getvalue()]),
        media_type="text/csv",
        headers={
            "Content-Disposition": f'attachment; filename="transactions-{start}-{end}.csv"'
        },
    )


@router.get("/export/summary.csv")
def export_simple_csv(
    start: date | None = None,
    end: date | None = None,
    session: Session = Depends(get_session),
) -> StreamingResponse:
    """One row per transaction, for spreadsheets. Plan section 10's "simple
    interoperability" case.

    This view is lossy by construction: a transaction split across two categories
    collapses to one row, and its `amount` is the net movement across liquid
    accounts, which is zero for a card purchase. transactions.csv is the
    canonical export; this one is for pasting into a spreadsheet.
    """
    today = clock_today(session)
    start = start or date(today.year, 1, 1)
    end = end or today

    liquid = {
        a.id for a in session.scalars(select(Account)) if a.kind in LIQUID_KINDS
    }
    grouped: dict[str, dict] = {}
    for txn, posting, account, category in _rows(session, start, end):
        row = grouped.setdefault(
            str(txn.id),
            {
                "booking_date": txn.booking_date.isoformat(),
                "description": txn.description,
                "merchant": txn.merchant or "",
                "categories": set(),
                "cash_amount": Decimal("0"),
                "expense_amount": Decimal("0"),
            },
        )
        if category is not None:
            row["categories"].add(category.name)
        if posting.account_id in liquid:
            row["cash_amount"] += posting.amount
        if account.kind == AccountKind.EXPENSE:
            row["expense_amount"] += posting.amount

    buffer = io.StringIO()
    writer = csv.writer(buffer)
    writer.writerow(
        ["booking_date", "description", "merchant", "categories",
         "cash_amount", "expense_amount", "currency"]
    )
    for row in sorted(grouped.values(), key=lambda r: r["booking_date"]):
        writer.writerow([
            row["booking_date"],
            row["description"],
            row["merchant"],
            "; ".join(sorted(row["categories"])),
            str(row["cash_amount"]),
            str(row["expense_amount"]),
            "GBP",
        ])

    buffer.seek(0)
    return StreamingResponse(
        iter([buffer.getvalue()]),
        media_type="text/csv",
        headers={
            "Content-Disposition": f'attachment; filename="summary-{start}-{end}.csv"'
        },
    )


@router.get("/export/backup.json")
def export_json(session: Session = Depends(get_session)) -> StreamingResponse:
    """Full-fidelity machine-readable backup of the ledger.

    The payload comes from `domain/backup.py`, which is also what a scheduled
    backup writes (B-A). Two serialisations would mean restore had only ever
    been tried against one of them.
    """
    return StreamingResponse(
        iter([backup_module.serialise(session)]),
        media_type="application/json",
        headers={"Content-Disposition": 'attachment; filename="backup.json"'},
    )


# --------------------------------------------------------------------------
# Restore (plan section 14)
# --------------------------------------------------------------------------


class RestoreResultOut(BaseModel):
    accounts: int
    categories: int
    transactions: int
    postings: int


class RestoreStatusOut(BaseModel):
    empty: bool


@router.get("/restore/status", response_model=RestoreStatusOut)
def restore_status(session: Session = Depends(get_session)) -> RestoreStatusOut:
    """Authoritative overwrite status; a paginated transaction list is not one."""
    return RestoreStatusOut(empty=restore_module.is_empty(session))


@router.post("/restore", response_model=RestoreResultOut)
async def restore_backup(
    payload: dict,
    replace: bool = False,
    session: Session = Depends(get_session),
) -> RestoreResultOut:
    """Rebuild the ledger from a backup produced by /export/backup.json.

    Refuses to overwrite a non-empty database unless `replace` is set: running a
    restore against the wrong database is the usual way to lose data with one of
    these. Validation happens before any write, so a malformed file leaves
    everything untouched.
    """
    try:
        result = restore_module.restore(session, payload, replace=replace)
    except restore_module.RestoreError as exc:
        session.rollback()
        raise HTTPException(status_code=422, detail=str(exc)) from exc
    return RestoreResultOut(
        accounts=result.accounts,
        categories=result.categories,
        transactions=result.transactions,
        postings=result.postings,
    )


# --------------------------------------------------------------------------
# XLSX and PDF (plan section 10's remaining two formats)
# --------------------------------------------------------------------------


def _period(session: Session, start: date | None, end: date | None) -> tuple[date, date]:
    today = clock_today(session)
    return start or date(today.year, 1, 1), end or today


@router.get("/export/transactions.xlsx")
def export_xlsx(
    start: date | None = None,
    end: date | None = None,
    session: Session = Depends(get_session),
) -> StreamingResponse:
    """Workbook with a posting sheet and a category summary sheet.

    Amounts are written as numbers rather than strings, because a spreadsheet
    column you cannot sum is not worth exporting. That costs exactness: a
    workbook stores IEEE doubles, so `transactions.csv` -- where every amount is
    a decimal string -- stays the canonical export, exactly as it is for
    `summary.csv`.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter

    start, end = _period(session, start, end)
    book = Workbook()

    postings = book.active
    postings.title = "Postings"
    headers = ["Date", "Description", "Merchant", "Account", "Kind",
               "Category", "Amount", "Currency"]
    postings.append(headers)

    for txn, posting, account, category in _rows(session, start, end):
        postings.append([
            txn.booking_date,
            txn.description,
            txn.merchant or "",
            account.name,
            account.kind.value,
            category.name if category else "",
            float(posting.amount),
            posting.currency,
        ])

    summary = book.create_sheet("By category")
    summary.append(["Category", "Amount"])
    period = analytics.summarise(session, start, end)
    for row in period.by_category:
        summary.append([row.name, float(row.amount)])
    summary.append([])
    summary.append(["Income", float(period.income)])
    summary.append(["Spending", float(period.expense)])
    summary.append(["Set aside", float(period.saved)])
    summary.append(["Net", float(period.net)])

    for sheet, money_cols, widths in (
        (postings, ("G",), (12, 34, 20, 20, 12, 22, 14, 10)),
        (summary, ("B",), (28, 14)),
    ):
        for cell in sheet[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="left")
        for index, width in enumerate(widths, start=1):
            sheet.column_dimensions[get_column_letter(index)].width = width
        for column in money_cols:
            for cell in sheet[column][1:]:
                cell.number_format = '#,##0.00'
        sheet.freeze_panes = "A2"
    for cell in postings["A"][1:]:
        cell.number_format = "yyyy-mm-dd"

    buffer = io.BytesIO()
    book.save(buffer)
    buffer.seek(0)
    return StreamingResponse(
        iter([buffer.getvalue()]),
        media_type=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
        headers={
            "Content-Disposition": (
                f'attachment; filename="transactions-{start}-{end}.xlsx"'
            )
        },
    )


@router.get("/export/statement.pdf")
def export_pdf(
    start: date | None = None,
    end: date | None = None,
    session: Session = Depends(get_session),
) -> StreamingResponse:
    """A statement for reading, not for re-importing.

    Every other export is a data interchange format; this one is the archival
    copy -- the thing that still means something opened in ten years with no
    application to load it into. So it states the period, the totals and the
    category breakdown, and stops there rather than dumping every posting.
    """
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.platypus import (
        Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle,
    )

    start, end = _period(session, start, end)
    period = analytics.summarise(session, start, end)

    styles = getSampleStyleSheet()
    muted = ParagraphStyle("muted", parent=styles["Normal"],
                           fontSize=9, textColor=colors.HexColor("#6b6a66"))

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=A4,
        leftMargin=20 * mm, rightMargin=20 * mm,
        topMargin=18 * mm, bottomMargin=18 * mm,
        title=f"Statement {start} to {end}", author="Personal Finance OS",
    )

    def money(value: Decimal) -> str:
        return f"{'-' if value < 0 else ''}£{abs(value):,.2f}"

    story = [
        Paragraph("Statement", styles["Title"]),
        Paragraph(f"{start:%-d %B %Y} to {end:%-d %B %Y}", muted),
        Spacer(1, 8 * mm),
    ]

    rate = (
        f"{period.savings_rate:.1%}" if period.savings_rate is not None else "—"
    )
    set_aside = (
        f"{period.set_aside_rate:.1%}" if period.set_aside_rate is not None else "—"
    )
    totals = Table(
        [
            ["Income", money(period.income)],
            ["Spending", money(period.expense)],
            ["Set aside", money(period.saved)],
            ["Net", money(period.net)],
            ["Savings rate", rate],
            ["Set-aside rate", set_aside],
        ],
        colWidths=[70 * mm, 40 * mm],
        hAlign="LEFT",
    )
    totals.setStyle(TableStyle([
        ("ALIGN", (1, 0), (1, -1), "RIGHT"),
        ("FONTNAME", (0, 3), (-1, 3), "Helvetica-Bold"),
        ("LINEBELOW", (0, 0), (-1, -2), 0.4, colors.HexColor("#dededa")),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
    ]))
    story += [totals, Spacer(1, 10 * mm)]

    if period.by_category:
        story.append(Paragraph("Spending by category", styles["Heading3"]))
        story.append(Spacer(1, 2 * mm))
        rows = [["Category", "Amount", "Share"]]
        total = sum((c.amount for c in period.by_category), Decimal("0"))
        for c in period.by_category:
            share = f"{c.amount / total:.1%}" if total else "—"
            rows.append([c.name, money(c.amount), share])
        table = Table(rows, colWidths=[95 * mm, 40 * mm, 25 * mm], hAlign="LEFT",
                      repeatRows=1)
        table.setStyle(TableStyle([
            ("ALIGN", (1, 0), (-1, -1), "RIGHT"),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("LINEBELOW", (0, 0), (-1, 0), 0.6, colors.HexColor("#9a9992")),
            ("LINEBELOW", (0, 1), (-1, -2), 0.3, colors.HexColor("#e8e8e4")),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ]))
        story.append(table)

    story += [
        Spacer(1, 12 * mm),
        Paragraph(
            "Figures are derived from postings at the time of export. "
            "Voided transactions are excluded. For a machine-readable copy use "
            "transactions.csv, where amounts are exact decimal strings.",
            muted,
        ),
    ]

    doc.build(story)
    buffer.seek(0)
    return StreamingResponse(
        iter([buffer.getvalue()]),
        media_type="application/pdf",
        headers={
            "Content-Disposition": f'attachment; filename="statement-{start}-{end}.pdf"'
        },
    )
