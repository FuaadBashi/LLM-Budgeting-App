"""XLSX and PDF export. Plan section 10's remaining two formats.

The tests that matter here are the ones asserting the new formats agree with the
CSV that was already reconciled to the ledger. A second export format is a second
chance to invent a number.
"""

from __future__ import annotations

import csv
import io
from datetime import date
from decimal import Decimal

import pytest
from fastapi.testclient import TestClient

from app.db import get_session
from app.domain import analytics
from app.main import app
from tests.conftest import post

START = date(2026, 8, 1)
END = date(2026, 8, 31)


@pytest.fixture
def client(session):
    app.dependency_overrides[get_session] = lambda: session
    with TestClient(app) as c:
        yield c
    app.dependency_overrides.clear()


@pytest.fixture
def month(session, accounts, categories):
    post(session, date(2026, 8, 1), "Salary",
         [(accounts["current"], "2500"), (accounts["salary"], "-2500")])
    post(session, date(2026, 8, 4), "Tesco",
         [(accounts["current"], "-62.40"),
          (accounts["groceries"], "62.40", categories["groceries"])],
         merchant="Tesco")
    post(session, date(2026, 8, 12), "Dinner",
         [(accounts["current"], "-46.50"),
          (accounts["groceries"], "46.50", categories["restaurants"])],
         merchant="Dishoom")
    post(session, date(2026, 8, 2), "Rent",
         [(accounts["current"], "-1200"),
          (accounts["groceries"], "1200", categories["rent"])])
    return session


def _sheets(payload: bytes) -> dict:
    from openpyxl import load_workbook

    book = load_workbook(io.BytesIO(payload))
    return {name: list(book[name].values) for name in book.sheetnames}


def test_xlsx_is_a_real_workbook(client, month):
    r = client.get(f"/api/export/transactions.xlsx?start={START}&end={END}")
    assert r.status_code == 200
    assert r.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml"
    )
    # The magic bytes for a zip container, which is what xlsx is.
    assert r.content[:2] == b"PK"
    sheets = _sheets(r.content)
    assert set(sheets) == {"Postings", "By category"}


def test_xlsx_posting_rows_match_the_csv_row_for_row(client, month):
    """The two exports read the same ledger, so they must agree.

    This is the test that catches a future filter added to one and not the other.
    """
    xlsx = _sheets(
        client.get(f"/api/export/transactions.xlsx?start={START}&end={END}").content
    )["Postings"]
    text = client.get(f"/api/export/transactions.csv?start={START}&end={END}").text
    rows = list(csv.DictReader(io.StringIO(text)))

    assert len(xlsx) - 1 == len(rows)
    for sheet_row, csv_row in zip(xlsx[1:], rows):
        assert sheet_row[0].date().isoformat() == csv_row["booking_date"]
        assert sheet_row[1] == csv_row["description"]
        assert Decimal(str(sheet_row[6])) == Decimal(csv_row["amount"])


def test_xlsx_amounts_are_numbers_not_text(client, month):
    """A column you cannot sum is not worth exporting."""
    rows = _sheets(
        client.get(f"/api/export/transactions.xlsx?start={START}&end={END}").content
    )["Postings"][1:]
    assert rows
    assert all(isinstance(row[6], (int, float)) for row in rows)


def test_xlsx_category_sheet_agrees_with_the_analytics_endpoint(client, month, session):
    sheet = _sheets(
        client.get(f"/api/export/transactions.xlsx?start={START}&end={END}").content
    )["By category"]
    summary = analytics.summarise(session, START, END)

    named = {row[0]: row[1] for row in sheet[1:] if row[0]}
    for category in summary.by_category:
        assert Decimal(str(named[category.name])) == category.amount
    assert Decimal(str(named["Income"])) == summary.income
    assert Decimal(str(named["Net"])) == summary.net


def test_pdf_is_a_real_pdf(client, month):
    r = client.get(f"/api/export/statement.pdf?start={START}&end={END}")
    assert r.status_code == 200
    assert r.headers["content-type"] == "application/pdf"
    assert r.content.startswith(b"%PDF-")
    assert r.content.rstrip().endswith(b"%%EOF")


def test_pdf_states_the_period_and_the_totals(client, month, session):
    """The archival copy has to be readable without the application.

    Extracting text from a PDF is awkward enough that it is tempting to assert
    only on the status code, which would pass for a blank page.
    """
    pytest.importorskip("pypdf", reason="text extraction needs pypdf")
    import pypdf

    r = client.get(f"/api/export/statement.pdf?start={START}&end={END}")
    reader = pypdf.PdfReader(io.BytesIO(r.content))
    text = "\n".join(page.extract_text() for page in reader.pages)

    summary = analytics.summarise(session, START, END)
    assert "Statement" in text
    assert "August 2026" in text
    assert f"{summary.income:,.2f}" in text
    assert f"{summary.expense:,.2f}" in text
    for category in summary.by_category:
        assert category.name in text


def test_pdf_excludes_voided_transactions(client, month, session, accounts, categories):
    """Same exclusion as every other engine -- X4."""
    pytest.importorskip("pypdf")
    import pypdf

    from app.models import TransactionStatus

    before = analytics.summarise(session, START, END).expense
    txn = post(session, date(2026, 8, 14), "Mistake",
               [(accounts["current"], "-999.99"),
                (accounts["groceries"], "999.99", categories["groceries"])])
    txn.status = TransactionStatus.VOIDED
    session.commit()

    r = client.get(f"/api/export/statement.pdf?start={START}&end={END}")
    text = "\n".join(
        p.extract_text() for p in pypdf.PdfReader(io.BytesIO(r.content)).pages
    )
    assert "999.99" not in text
    assert analytics.summarise(session, START, END).expense == before

    # The workbook must agree; a second export is a second chance to diverge.
    sheet = _sheets(
        client.get(f"/api/export/transactions.xlsx?start={START}&end={END}").content
    )["Postings"]
    assert all("Mistake" not in str(row[1]) for row in sheet[1:])


def test_both_formats_default_to_the_year_so_far(client, month):
    """No date range must not mean no data."""
    assert client.get("/api/export/transactions.xlsx").status_code == 200
    assert client.get("/api/export/statement.pdf").status_code == 200


def test_empty_period_still_produces_valid_files(client, session):
    """An empty month is a legitimate answer, not an error."""
    xlsx = client.get("/api/export/transactions.xlsx?start=2020-01-01&end=2020-01-31")
    pdf = client.get("/api/export/statement.pdf?start=2020-01-01&end=2020-01-31")
    assert xlsx.status_code == 200 and xlsx.content[:2] == b"PK"
    assert pdf.status_code == 200 and pdf.content.startswith(b"%PDF-")
